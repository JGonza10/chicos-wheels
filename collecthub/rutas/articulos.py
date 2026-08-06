"""Inventario de piezas y su historial de precios."""
import datetime
import json

import openpyxl
from flask import Blueprint, g, jsonify, request

from ..db import bd, todos, transaccion, uno
from ..util import ErrorApp, entero, fecha, hoy, num, texto, uid

bp = Blueprint("articulos", __name__)

TIPOS = ("Hot Wheels", "Pokémon")
ESTATUS = ("Disponible", "En negociación", "Conservar")

# Mismo orden que scripts/generar_plantilla.py — si cambian los campos del
# inventario, hay que actualizar ambos.
COLUMNAS_PLANTILLA = (
    "tipo", "nombre", "numero", "anio", "cantidad", "precio_compra", "valor_estimado",
    "fecha_adq", "estatus", "fuente", "ubicacion", "codigo", "serie", "color",
    "expansion", "rareza", "grado", "cert", "sub", "estado", "notas",
)
MAX_FILAS_IMPORTACION = 2000


def limpiar(b: dict, parcial: bool = False) -> dict:
    """Normaliza lo que llega del cliente. Nunca confiamos en el navegador."""
    o = {}

    def poner(clave, valor):
        if not parcial or clave in b:
            o[clave] = valor

    if "tipo" in b or not parcial:
        if b.get("tipo") not in TIPOS:
            raise ErrorApp("El tipo debe ser Hot Wheels o Pokémon")
        o["tipo"] = b["tipo"]
    if "nombre" in b or not parcial:
        nombre = texto(b.get("nombre"), 160)
        if not nombre:
            raise ErrorApp("La pieza necesita un nombre")
        o["nombre"] = nombre

    poner("numero", texto(b.get("numero"), 40))
    poner("anio", entero(b.get("anio")) if b.get("anio") else None)
    poner("serie", texto(b.get("serie"), 60))
    poner("color", texto(b.get("color"), 60))
    poner("expansion", texto(b.get("expansion"), 80))
    poner("rareza", texto(b.get("rareza"), 60))
    poner("grado", texto(b.get("grado"), 40))
    poner("cert", texto(b.get("cert"), 40))
    poner("sub", texto(b.get("sub"), 60))
    poner("estado", texto(b.get("estado"), 60))
    poner("cantidad", max(0, entero(b.get("cantidad"), 1)))
    poner("estatus", b["estatus"] if b.get("estatus") in ESTATUS else "Disponible")
    poner("precio_compra", max(0.0, num(b.get("precio_compra"))))
    poner("valor_estimado", max(0.0, num(b.get("valor_estimado"))))
    poner("fecha_adq", fecha(b["fecha_adq"]) if b.get("fecha_adq") else "")
    poner("fuente", texto(b.get("fuente"), 60))
    poner("ubicacion", texto(b.get("ubicacion"), 80))
    poner("codigo", texto(b.get("codigo"), 60))
    poner("foto", texto(b.get("foto"), 500))
    poner("notas", texto(b.get("notas"), 1000))
    poner("grail", 1 if b.get("grail") else 0)
    checks = b.get("checks")
    poner("checks", json.dumps([entero(c) for c in checks] if isinstance(checks, list) else []))
    return o


def mio(id_art: str) -> dict:
    a = uno("SELECT * FROM v_articulos WHERE id=? AND usuario_id=?", (id_art, g.usuario_id))
    if not a:
        raise ErrorApp("Esa pieza no existe en tu inventario", 404)
    return a


@bp.get("")
def listar():
    return jsonify(todos(
        "SELECT * FROM v_articulos WHERE usuario_id=? ORDER BY creado_en DESC", (g.usuario_id,)))


@bp.get("/<id_art>")
def detalle(id_art):
    a = mio(id_art)
    a["valuaciones"] = todos(
        "SELECT * FROM valuaciones WHERE articulo_id=? ORDER BY fecha", (id_art,))
    return jsonify(a)


def _insertar_articulo(con, datos: dict) -> str:
    """Inserta un artículo ya validado por limpiar() y su valuación inicial.
    Reutilizada por crear() (un artículo) e importar() (un lote desde .xlsx)."""
    nuevo_id = uid("PKM" if datos["tipo"] == "Pokémon" else "HW")
    columnas = list(datos.keys())
    con.execute(
        f"INSERT INTO articulos (id,usuario_id,cant_inicial,{','.join(columnas)}) "
        f"VALUES (?,?,?,{','.join('?' * len(columnas))})",
        [nuevo_id, g.usuario_id, datos["cantidad"]] + [datos[c] for c in columnas])
    if datos["valor_estimado"] > 0:
        con.execute(
            "INSERT INTO valuaciones (id,articulo_id,fecha,valor,fuente) VALUES (?,?,?,?,?)",
            (uid("VAL"), nuevo_id, datos["fecha_adq"] or hoy(),
             datos["valor_estimado"], "Registro inicial"))
    return nuevo_id


@bp.post("")
def crear():
    datos = limpiar(request.get_json(silent=True) or {})
    with transaccion() as con:
        nuevo_id = _insertar_articulo(con, datos)
    return jsonify(uno("SELECT * FROM v_articulos WHERE id=?", (nuevo_id,))), 201


def _valor_celda(v):
    """openpyxl entrega fechas como datetime; el resto de los campos, tal cual
    la celda los tenga. Todo se vuelve texto/número simple antes de limpiar()."""
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    return v


@bp.post("/importar")
def importar():
    """Carga masiva desde el .xlsx de scripts/generar_plantilla.py. Cada fila
    se valida con la MISMA función que usa el alta manual (limpiar()); una
    fila con error no tumba las demás, se reporta y se sigue con el resto."""
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        raise ErrorApp("Adjunta el archivo .xlsx de la plantilla")
    if not archivo.filename.lower().endswith(".xlsx"):
        raise ErrorApp("El archivo debe ser .xlsx (usa la plantilla que se descarga aquí mismo)")

    try:
        libro = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    except Exception:
        raise ErrorApp("No se pudo leer el archivo. ¿Es un .xlsx válido y no está dañado?")

    hoja = libro["Inventario"] if "Inventario" in libro.sheetnames else libro.worksheets[0]
    filas = hoja.iter_rows(values_only=True)
    try:
        encabezado = [str(c).strip().lower() if c is not None else "" for c in next(filas)]
    except StopIteration:
        raise ErrorApp("El archivo está vacío")

    columnas_usadas = {nombre: i for i, nombre in enumerate(encabezado) if nombre in COLUMNAS_PLANTILLA}
    if "tipo" not in columnas_usadas or "nombre" not in columnas_usadas:
        raise ErrorApp("Faltan columnas obligatorias (tipo, nombre). ¿Es la plantilla original?")

    insertados, errores, n_fila = 0, [], 1
    with transaccion() as con:
        for cruda in filas:
            n_fila += 1
            if n_fila - 1 > MAX_FILAS_IMPORTACION:
                errores.append({"fila": n_fila, "mensaje": f"Se alcanzó el máximo de {MAX_FILAS_IMPORTACION} filas por archivo; el resto no se procesó."})
                break
            if cruda is None or all(c is None for c in cruda):
                continue  # fila en blanco, se ignora sin marcarla como error
            fila_dict = {nombre: _valor_celda(cruda[i]) for nombre, i in columnas_usadas.items()
                         if i < len(cruda) and cruda[i] is not None}
            if not fila_dict.get("nombre"):
                continue  # fila vacía salvo alguna celda suelta
            try:
                datos = limpiar(fila_dict)
                _insertar_articulo(con, datos)
                insertados += 1
            except ErrorApp as e:
                errores.append({"fila": n_fila, "mensaje": e.mensaje})
            except Exception:
                # No debe pasar (limpiar() ya valida todo), pero si pasa, que
                # se reporte esa fila y siga con las demás en vez de tirar
                # el lote entero (con transaccion() haría rollback de todo
                # si el error se dejara escapar del with).
                errores.append({"fila": n_fila, "mensaje": "No se pudo guardar esta fila."})

    return jsonify(insertados=insertados, errores=errores), 201


@bp.patch("/<id_art>")
def editar(id_art):
    actual = mio(id_art)
    datos = limpiar(request.get_json(silent=True) or {}, parcial=True)
    if not datos:
        return jsonify(actual)
    asigna = ",".join(f"{c}=?" for c in datos)
    bd().execute(
        f"UPDATE articulos SET {asigna}, actualizado_en=datetime('now') "
        f"WHERE id=? AND usuario_id=?",
        list(datos.values()) + [id_art, g.usuario_id])
    bd().commit()
    return jsonify(uno("SELECT * FROM v_articulos WHERE id=?", (id_art,)))


@bp.delete("/<id_art>")
def eliminar(id_art):
    mio(id_art)
    vigentes = uno("SELECT COUNT(*) n FROM apartados WHERE articulo_id=? AND estatus='Vigente'",
                   (id_art,))["n"]
    if vigentes:
        raise ErrorApp("Esta pieza tiene un apartado vigente. Cancélalo primero.", 409)
    bd().execute("DELETE FROM articulos WHERE id=?", (id_art,))
    bd().commit()
    return jsonify(ok=True)


# ---------- Historial de precios ----------

@bp.post("/<id_art>/valuaciones")
def agregar_valuacion(id_art):
    mio(id_art)
    datos = request.get_json(silent=True) or {}
    valor = num(datos.get("valor"))
    if valor <= 0:
        raise ErrorApp("Escribe el valor que observaste en el mercado")

    with transaccion() as con:
        con.execute("INSERT INTO valuaciones (id,articulo_id,fecha,valor,fuente) VALUES (?,?,?,?,?)",
                    (uid("VAL"), id_art, fecha(datos.get("fecha")), valor,
                     texto(datos.get("fuente"), 60)))
        if datos.get("actualizar") is not False:
            con.execute(
                "UPDATE articulos SET valor_estimado=?, actualizado_en=datetime('now') WHERE id=?",
                (valor, id_art))

    return jsonify(
        articulo=uno("SELECT * FROM v_articulos WHERE id=?", (id_art,)),
        valuaciones=todos("SELECT * FROM valuaciones WHERE articulo_id=? ORDER BY fecha", (id_art,)),
    ), 201


@bp.delete("/<id_art>/valuaciones/<id_val>")
def borrar_valuacion(id_art, id_val):
    mio(id_art)
    bd().execute("DELETE FROM valuaciones WHERE id=? AND articulo_id=?", (id_val, id_art))
    bd().commit()
    return jsonify(ok=True)
