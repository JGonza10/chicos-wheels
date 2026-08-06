#!/usr/bin/env python3
"""
Genera static/plantilla-inventario.xlsx: la plantilla que se descarga desde
Datos → Importar inventario y que el botón de "Importar" sabe leer de vuelta.

Los encabezados son EXACTAMENTE los nombres de campo que espera el backend
(los mismos que valida collecthub/rutas/articulos.py → limpiar()), para que
no haga falta traducir nombres de columna al importar. La hoja "Instrucciones"
explica qué significa cada uno en español llano.

Se corre a mano cada vez que cambian los campos del inventario:

    python scripts/generar_plantilla.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
DESTINO = RAIZ / "static" / "plantilla-inventario.xlsx"

# (columna, ancho, ¿es texto forzado? para no perder ceros/diagonales como "004/102")
COLUMNAS = [
    ("tipo", 14, False),
    ("nombre", 32, False),
    ("numero", 12, True),
    ("anio", 8, False),
    ("cantidad", 10, False),
    ("precio_compra", 14, False),
    ("valor_estimado", 14, False),
    ("fecha_adq", 13, False),
    ("estatus", 16, False),
    ("fuente", 20, False),
    ("ubicacion", 16, False),
    ("codigo", 14, True),
    ("serie", 22, False),
    ("color", 14, False),
    ("expansion", 22, False),
    ("rareza", 18, False),
    ("grado", 10, False),
    ("cert", 14, True),
    ("sub", 16, False),
    ("estado", 16, False),
    ("notas", 40, False),
]

EJEMPLOS = [
    {"tipo": "Hot Wheels", "nombre": "Nissan Skyline GT-R (BNR34)", "numero": "088/250", "anio": 2024,
     "cantidad": 1, "precio_compra": 45, "valor_estimado": 120, "fecha_adq": "2026-06-01",
     "estatus": "Disponible", "fuente": "Tienda / retail", "ubicacion": "Caja A", "codigo": "",
     "serie": "Mainline", "color": "Azul", "sub": "", "estado": "Nuevo, blíster sellado",
     "notas": "Ejemplo — bórrame antes de importar"},
    {"tipo": "Pokémon", "nombre": "Charizard ex", "numero": "205/165", "anio": 2023,
     "cantidad": 1, "precio_compra": 800, "valor_estimado": 1420, "fecha_adq": "2026-05-15",
     "estatus": "Conservar", "fuente": "Convención", "ubicacion": "Caja fuerte", "codigo": "",
     "expansion": "151", "rareza": "Special Illustration Rare", "grado": "", "cert": "",
     "sub": "", "estado": "Mint", "notas": "Ejemplo — bórrame antes de importar"},
]

TIPOS = ["Hot Wheels", "Pokémon"]
ESTATUS = ["Disponible", "En negociación", "Conservar"]

AZUL = "0A1120"
DORADO = "FFD84D"


def hoja_inventario(wb):
    ws = wb.active
    ws.title = "Inventario"

    encabezado = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor=AZUL)

    for i, (nombre, ancho, _es_texto) in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=i, value=nombre)
        celda.font = encabezado
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center")
        ws.column_dimensions[celda.column_letter].width = ancho

    for fila, ejemplo in enumerate(EJEMPLOS, start=2):
        for i, (nombre, _ancho, es_texto) in enumerate(COLUMNAS, start=1):
            valor = ejemplo.get(nombre, "")
            celda = ws.cell(row=fila, column=i, value=valor)
            if es_texto:
                celda.number_format = "@"

    # Fuerza texto en columnas donde un cero a la izquierda o una diagonal
    # ("004/102") no debe convertirse en número/fecha por accidente.
    for i, (_nombre, _ancho, es_texto) in enumerate(COLUMNAS, start=1):
        if es_texto:
            for fila in range(2, 500):
                ws.cell(row=fila, column=i).number_format = "@"

    # Menús desplegables para los dos campos con valores fijos.
    col_tipo = next(i for i, (n, *_r) in enumerate(COLUMNAS, start=1) if n == "tipo")
    col_estatus = next(i for i, (n, *_r) in enumerate(COLUMNAS, start=1) if n == "estatus")
    dv_tipo = DataValidation(type="list", formula1=f'"{",".join(TIPOS)}"', allow_blank=False,
                              showErrorMessage=True, errorTitle="Tipo inválido",
                              error="Escribe exactamente Hot Wheels o Pokémon.")
    dv_estatus = DataValidation(type="list", formula1=f'"{",".join(ESTATUS)}"', allow_blank=True,
                                 showErrorMessage=True, errorTitle="Estatus inválido",
                                 error="Elige uno de la lista.")
    ws.add_data_validation(dv_tipo)
    ws.add_data_validation(dv_estatus)
    letra_tipo = ws.cell(row=1, column=col_tipo).column_letter
    letra_estatus = ws.cell(row=1, column=col_estatus).column_letter
    dv_tipo.add(f"{letra_tipo}2:{letra_tipo}500")
    dv_estatus.add(f"{letra_estatus}2:{letra_estatus}500")

    ws.freeze_panes = "A2"


def hoja_instrucciones(wb):
    ws = wb.create_sheet("Instrucciones")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90

    titulo = ws.cell(row=1, column=1, value="Cómo llenar esta plantilla")
    titulo.font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:B1")

    filas = [
        ("tipo", "Obligatorio. Exactamente 'Hot Wheels' o 'Pokémon' (hay un menú desplegable en la columna)."),
        ("nombre", "Obligatorio. El nombre del modelo o de la carta."),
        ("numero", "Número de colección. Se guarda como texto, así que '004/102' no se convierte en fecha."),
        ("anio", "Año de fabricación o de la expansión. Solo el número, ej. 2024."),
        ("cantidad", "Cuántas piezas idénticas tienes. Si lo dejas vacío, se asume 1."),
        ("precio_compra", "Lo que pagaste por unidad, sin símbolo de moneda (ej. 45.50)."),
        ("valor_estimado", "Lo que crees que vale hoy en el mercado, sin símbolo de moneda."),
        ("fecha_adq", "Fecha en que la adquiriste, formato AAAA-MM-DD (ej. 2026-06-01). Si la dejas vacía, se usa hoy."),
        ("estatus", "Disponible, En negociación o Conservar (hay un menú desplegable). Si lo dejas vacío, se asume Disponible."),
        ("fuente", "Dónde la conseguiste, ej. 'Tienda / retail', 'Bazar o tianguis', 'Convención'."),
        ("ubicacion", "Dónde la guardas físicamente, ej. 'Caja A'."),
        ("codigo", "Tu código interno de inventario o de barras, si usas uno."),
        ("serie", "Solo Hot Wheels: Mainline, Treasure Hunt, Super Treasure Hunt, etc."),
        ("color", "Solo Hot Wheels: color de la carrocería."),
        ("expansion", "Solo Pokémon: nombre de la expansión/set."),
        ("rareza", "Solo Pokémon: Common, Rare, Holo Rare, Secret Rare, etc."),
        ("grado", "Solo Pokémon: si está graduada, el grado (ej. PSA 9)."),
        ("cert", "Solo Pokémon: número de certificado de graduación, si aplica."),
        ("sub", "Sub-línea o variante, si aplica a cualquiera de los dos tipos."),
        ("estado", "Estado físico de la pieza en tus palabras (ej. 'Nuevo, blíster sellado')."),
        ("notas", "Cualquier observación libre."),
    ]
    for i, (campo, explicacion) in enumerate(filas, start=3):
        c1 = ws.cell(row=i, column=1, value=campo)
        c1.font = Font(bold=True)
        ws.cell(row=i, column=2, value=explicacion).alignment = Alignment(wrap_text=True, vertical="top")

    aviso = ws.cell(row=len(filas) + 4, column=1,
                     value="Borra las dos filas de ejemplo de la hoja \"Inventario\" antes de importar tu propio inventario.")
    aviso.font = Font(italic=True, color="A6192E")
    ws.merge_cells(start_row=len(filas) + 4, start_column=1, end_row=len(filas) + 4, end_column=2)
    ws.cell(row=len(filas) + 4, column=1).alignment = Alignment(wrap_text=True)


def main():
    wb = Workbook()
    hoja_inventario(wb)
    hoja_instrucciones(wb)
    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINO)
    print(f"Plantilla generada en {DESTINO}")


if __name__ == "__main__":
    main()
